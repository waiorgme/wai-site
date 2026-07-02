#!/usr/bin/env python3
"""Claim-wave import: cleaned member list (xlsx) -> Convex importedMembers.

Usage (Issam runs this; Claude is code-only on the deployment):

    # counts only, touches nothing (the default):
    python3 scripts/import-members.py \
        "/Users/ismac/Documents/Projects/WAI/03 Members/WAI-ME Member List (Cleaned) 2026-06-13.xlsx"

    # actually push to the configured Convex deployment:
    python3 scripts/import-members.py "<same path>" --live

What it does (spec: specs/claim-wave.spec.md):
- reads the "Members (Cleaned)" sheet
- normalises each row (lower-cased email, WAIME-### -> number, birthday -> ISO)
- marks anyone whose birthday makes them under 18 TODAY as suppressed_minor
  (the under-18 hold-back; guardian flow is a later slice)
- pushes idempotent batches of 100 to internal.importedMembers.importBatch
  via `npx convex run` (re-running never duplicates)
- finally raises the membership-number counter floor above the highest legacy
  number (DATA-1) via internal.importedMembers.raiseCounterFloor

The member list itself must NEVER be committed to this repo (PII).
Requires: python3 + openpyxl, and a configured Convex deployment (.env.local).
"""

import datetime as dt
import hashlib
import json
import re
import subprocess
import sys
from pathlib import Path

BATCH = 100
SHEET = "Members (Cleaned)"


def clean(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def parse_legacy_number(raw):
    if not raw:
        return None
    m = re.match(r"^WAIME[\s-]*(\d{1,6})$", str(raw).strip(), re.IGNORECASE)
    return int(m.group(1)) if m else None


def normalize_birthday(raw):
    if raw is None:
        return None
    if isinstance(raw, dt.datetime) or isinstance(raw, dt.date):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()[:10]
    try:
        parsed = dt.date.fromisoformat(s)
    except ValueError:
        return None
    if parsed.year < 1900 or parsed > dt.date.today():
        return None
    return s


def age_years(dob_iso, today):
    d = dt.date.fromisoformat(dob_iso)
    years = today.year - d.year
    if (today.month, today.day) < (d.month, d.day):
        years -= 1
    return years


def created_at_iso(raw):
    if isinstance(raw, (dt.datetime, dt.date)):
        return raw.strftime("%Y-%m-%d")
    s = clean(raw)
    return s[:10] if s else None


def main():
    args = [a for a in sys.argv[1:] if a not in ("--dry-run", "--live")]
    # Safe by default: pushing to the deployment is an explicit opt-in.
    dry_run = "--live" not in sys.argv
    if len(args) != 1:
        sys.exit(__doc__)
    xlsx = Path(args[0])
    if not xlsx.exists():
        sys.exit(f"Not found: {xlsx}")

    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        sys.exit("openpyxl missing: pip3 install openpyxl")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f"Sheet '{SHEET}' not found; sheets: {wb.sheetnames}")
    ws = wb[SHEET]
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows_iter)]
    idx = {name: i for i, name in enumerate(header)}
    required = ["email", "first_name", "last_name", "membership_number"]
    for col in required:
        if col not in idx:
            sys.exit(f"Column '{col}' missing; header: {header}")

    today = dt.date.today()
    out, skipped, blockers, seen_emails, seen_row_ids = [], [], [], set(), set()
    max_legacy = 0

    for n, r in enumerate(rows_iter, start=2):
        get = lambda col: r[idx[col]] if col in idx and idx[col] < len(r) else None
        email = clean(get("email"))
        if not email or "@" not in email:
            skipped.append((n, "no usable email"))
            continue
        email = email.lower()
        if email in seen_emails:
            # Two people, one email = the Stage 0 conflict model. Silently
            # skipping would let the email holder claim whichever row the
            # backend saw first, so a duplicate BLOCKS the import until a
            # human resolves it (known case: cleaned-list row 646).
            blockers.append((n, f"duplicate email {email}; resolve by hand before import"))
            continue
        seen_emails.add(email)

        first = clean(get("first_name")) or ""
        last = clean(get("last_name")) or ""
        name = f"{first} {last}".strip()
        if not name:
            skipped.append((n, "no name"))
            continue

        legacy_number = parse_legacy_number(get("membership_number"))
        if not legacy_number:
            # Every claimable migrated row must carry its legacy WAIME number:
            # the certificate promise is her ORIGINAL number, never a fresh
            # counter value (DATA-1). A missing/unparseable number blocks the
            # import until a human fixes the sheet.
            blockers.append((n, f"missing or unparseable membership_number for {email}"))
            continue
        max_legacy = max(max_legacy, legacy_number)
        dob = normalize_birthday(get("birthday"))
        gender = clean(get("gender"))
        gender = gender.lower() if gender and gender.lower() in ("female", "male") else None

        # STABLE row identity (never a sheet position, which shifts when rows
        # are added/removed): the legacy membership number when present,
        # otherwise the original email. Re-imports with corrected emails
        # update the same row instead of duplicating the member.
        row_id = f"waime:{legacy_number}" if legacy_number else f"email:{email}"
        if row_id in seen_row_ids:
            skipped.append((n, f"duplicate legacy id {row_id}; resolve by hand"))
            continue
        seen_row_ids.add(row_id)
        row = {
            "legacy_row_id": row_id,
            "normalized_email": email,
            "name": name,
            "mobile": clean(get("mobile")),
            "dob_if_known": dob,
            "legacy_position": clean(get("position")),
            "legacy_company": clean(get("company")),
            "legacy_bio": clean(get("bio")),
            "gender": gender,
            "nationality": clean(get("nationality")),
            "country_of_residence": clean(get("country_residency")),
            "legacy_membership_number": legacy_number,
            "legacy_created_at": created_at_iso(get("created_at")),
            "suppressed_minor": bool(dob and age_years(dob, today) < 18),
        }
        # Hash the content (not the row id) so edits re-flow on re-import.
        row["legacy_row_hash"] = hashlib.sha256(
            json.dumps({k: v for k, v in row.items() if k != "legacy_row_id"},
                       sort_keys=True, default=str).encode()
        ).hexdigest()[:32]
        # Convex validators reject nulls for optional fields; drop them.
        out.append({k: v for k, v in row.items() if v is not None})

    minors = sum(1 for r in out if r.get("suppressed_minor"))
    print(f"rows ready: {len(out)}  suppressed minors: {minors}  "
          f"max legacy number: {max_legacy}  skipped: {len(skipped)}")
    for n, why in skipped[:10]:
        print(f"  skipped row {n}: {why}")

    if blockers:
        print(f"IMPORT BLOCKED: {len(blockers)} row(s) need a human before ANY import:")
        for n, why in blockers:
            print(f"  row {n}: {why}")
        if dry_run:
            print("dry run: nothing would be sent while blockers exist")
            return
        sys.exit(1)

    if dry_run:
        print("dry run: nothing sent")
        return

    repo = Path(__file__).resolve().parent.parent
    for i in range(0, len(out), BATCH):
        batch = out[i:i + BATCH]
        result = subprocess.run(
            ["npx", "convex", "run", "importedMembers:importBatch",
             json.dumps({"rows": batch})],
            cwd=repo, capture_output=True, text=True,
        )
        if result.returncode != 0:
            sys.exit(f"batch {i // BATCH + 1} failed:\n{result.stderr}")
        print(f"batch {i // BATCH + 1}: {result.stdout.strip()}")

    result = subprocess.run(
        ["npx", "convex", "run", "importedMembers:raiseCounterFloor",
         json.dumps({"maxLegacyNumber": max_legacy})],
        cwd=repo, capture_output=True, text=True,
    )
    if result.returncode != 0:
        sys.exit(f"counter floor failed:\n{result.stderr}")
    print(f"counter floor: {result.stdout.strip()}")


if __name__ == "__main__":
    main()
